# -*- coding: utf-8 -*-
# by Shawn Gu

import pandas as pd
import numpy as np
import openpyxl as ox
import datetime,re
#__________________________________ MySQL setting ___________________________________

from sqlalchemy import create_engine,Table,Column,Integer,String,MetaData,ForeignKey
engine=create_engine('mysql+pymysql://root:root@localhost:3306/test?charset=utf8',echo=True,encoding='utf-8')

#_____________________________________________________________________________________
this_term = '20151231'   # 本期年
this_term_out_range = "2015*"   # 本期年, 抛出正常range的列
last_term = '20141231'   # 上期年

lap0=datetime.datetime.now()

df = pd.read_excel('D:/pyscripts/FS.xlsx', sheetname='FS', header=None)    # 读取数据
df.dropna(axis=1, how='all')                                             # dropna 列

if df.shape[1] == 8:
    df.columns = ["Ticker", "Attribute_1", "Account", "Attribute_2", last_term, this_term, this_term_out_range, 'redundant']   # 设置每列 column 名称
elif df.shape[1] == 7:
    df.columns = ["Ticker", "Attribute_1", "Account", "Attribute_2", last_term, this_term, 'redundant']

df['Attribute_2'] = df['Attribute_2'].fillna(method='ffill')              # ffill Attribute_2 的 NaN value
data_frame = df.set_index(['Ticker', 'Attribute_1', 'Attribute_2', 'Account','redundant'])    # 将部分所需column设置成需要顺序的index
data_frame = data_frame.dropna(axis=0, how='all').reset_index().fillna(0).replace(to_replace='-', value=0)            # dropna 行, reset_index, 把剩下的NaN替换成 0
data_frame[this_term]=data_frame[this_term_out_range]+data_frame[this_term]                    # 2015年 = 2015* 值
data_frame = data_frame.drop(axis=1,labels=this_term_out_range).drop(axis=1,labels='redundant').replace(to_replace=0,value=np.nan)                  # drop 2015*, 替换剩下的 0 为NaN

data_frame.to_excel('del.xlsx', index=False)                  # 输出 excel

lap1=datetime.datetime.now()

print('*** Pandas Process Done: ',lap1-lap0,' ***')

wb_orig=ox.load_workbook('D:/pyscripts/extract_FS/del.xlsx')
ws_orig=wb_orig.active
wb_dest=ox.load_workbook('C:/Users/Shawn Gu/Desktop/Copy of 上海时银资产管理有限公司-报表表格设计.xlsx')
tickers=set([ws_orig.cell(row=i, column=1).value for i in range(2,ws_orig.max_row)])
# ______________________________________________________________________________________________________________________
for i in range(1,ws_orig.max_row):
    if ws_orig.cell(row=i, column=3).value=='流动负债' and ws_orig.cell(row=i, column=4).value=='递延收益':  # 更改 短期递延收益 名称
        ws_orig.cell(row=i, column=4).value='短期递延收益'
wb_orig.save('D:/pyscripts/extract_FS/del.xlsx')
# ______________________________________________________________________________________________________________________
col_for_ticker_date=6
for t in tickers:    # 先将ticker和 日期 填入 数据库上的title, 为了下面可以对应
    for d in [this_term,last_term]: #上期与本期的日期
        for sheet in ['BS_consolidated','BS_company', 'IS_consolidated','IS_company','CF_consolidated','CF_company']:
            ws_dest=wb_dest.get_sheet_by_name(sheet)
            ws_dest.cell(row=2,column=col_for_ticker_date).value=t
            ws_dest.cell(row=3,column=col_for_ticker_date).value=d
            for num in range(6, ws_dest.max_column + 1):
                if re.search('consolidated', sheet):
                    ws_dest.cell(row=4, column=num).value = 2       # 2. 合并未调整
                elif re.search('company', sheet):
                    ws_dest.cell(row=4, column=num).value = 4       # 4. 母公司未调整
                ws_dest.cell(row=5, column=num).value = 0           # 0. 一般企业
                ws_dest.cell(row=1, column=num).value = str(ws_dest.cell(row=2, column=num).value)[:6] +\
                                                        str(ws_dest.cell(row=3, column=num).value) +\
                                                        str(ws_dest.cell(row=4, column=num).value) +\
                                                        str(ws_dest.cell(row=5, column=num).value)      # primary key = ticker(row2) + yr(row3) + mark(row4) + enterprise(row5)
        col_for_ticker_date +=1

for r in range(1,ws_orig.max_row):
    for i in [('合并资产负债表','BS_consolidated'),('母公司资产负债表','BS_company'),('合并利润表','IS_consolidated'),('母公司利润表','IS_company'),('合并现金流量表','CF_consolidated'),('母公司现金流量表','CF_company')]:
        if ws_orig.cell(row=r, column=2).value == i[0]: # 根据 orig 表里的项目和 i 对应
            ws_dest = wb_dest.get_sheet_by_name(i[1])   # 当与列表中的值相对应时, 那么就可以去对应 sheet 中写入数据
            for c in range(6,ws_dest.max_column+1):
                if ws_orig.cell(row=r, column=1).value==ws_dest.cell(row=2,column=c).value:  # 对应 ticker
                    for x in range(2,ws_dest.max_row):  # 对应完 ticker 后,将其往下写
                        if ws_orig.cell(row=r,column=4).value==ws_dest.cell(row=x,column=2).value:
                            if ws_dest.cell(row=3,column=c).value==this_term:   # 对应 年份
                                c1=6    # 若是本期, 就取 orig  第6列写入
                            elif ws_dest.cell(row=3,column=c).value==last_term:
                                c1=5    # 若是上期, 就取 orig  第5列写入
                            ws_dest.cell(row=x,column=c).value = ws_orig.cell(row=r,column=c1).value
                            break       # 每次写一行, 即break
lap2=datetime.datetime.now()

wb_dest.save('del2.xlsx')
for sheet in ['BS_consolidated','BS_company', 'IS_consolidated','IS_company','CF_consolidated','CF_company']:
    df=pd.read_excel('D:/pyscripts/extract_FS/del2.xlsx',sheetname=sheet,header=None)
    print('pd ising')
    df=df.drop(axis=1,labels=[0,1,2,3,4,5]).T
    print('T ing')
    df.to_excel(sheet+'.xlsx',header=False,index=False)
    print('exporting')
    # if sheet=='BS_consolidated' or sheet=='BS_company':
    #     df.to_sql('bs',con=engine,if_exists='append', index=False, chunksize=1000)
    # else:
    #     continue

print('*** ALL DONE: ',lap2-lap1,' ***')