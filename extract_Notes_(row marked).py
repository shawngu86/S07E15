# -*- coding: utf-8 -*-
# by Shawn Gu

import openpyxl as ox
import datetime
import os
import re
import numpy as np
import pandas as pd
from openpyxl.styles import Font,Color,colors

# _________________________ create a workbook for data storage_______________________

wb1 = ox.load_workbook('D:/pyscripts/FS.xlsx', data_only=True)  # 打开要录入的目的文件
wb_LTA=ox.load_workbook('D:/pyscripts/LTA.xlsx', data_only=True)
ws_LTA=wb_LTA.active

# ___________________________________________________________________________________
def clean_list(List):
    List = [str(List[i]).replace(" ", "") for i in range(len(List))]        # list value del space
    List = [str(List[i]).replace("\n", "") for i in range(len(List))]       # list value del \n
    List = [str(List[i]).replace("None", "") for i in range(len(List))]      # list value del "None"
    tmp = []
    for i in range(len(List)):
        try:
            tmp.append(float(List[i]))
        except ValueError:
            tmp.append(List[i])
    List = tmp
    return List


def trim_items_a(old, in_which):
    return re.match(old.encode('utf-8'), in_which.encode('utf-8')) is not None


# ___________________________________________________________________________________
row = []
LTA = []   # 储存LTA
path = "C:/Users/Shawn Gu/Desktop/CN_annual"             # files need to be extracted

dic = {'(\S*|\S*\s*)成本构成\S*': '成本分析', '(\S*|\S*\s*)费用按性质分类\S*': '成本分析', '(\S*|\S*\s*)成本分析\S*': '成本分析', '(\S*|\S*\s*)股份变动\S*': '股份变动',
       '(\S*|\S*\s*)研发投入\S*': '研发投入', '(\S*|\S*\s*)成本进行计量\S*': '折旧方法', '(\S*|\S*\s*|\S*\s*\S*)折旧方法\S*': '折旧方法',
       '(\S*|\S*\s*)存货跌价\S*': '存货跌价准备', '(\S*|\S*\s*)计价方法\S*': '折旧方法', '(\S*|\S*\s*)本期计提\S*': '本期计提',
       '(\S*|\S*\s*)货币资金\S*': '货币资金', '(\S*|\S*\s*)以公允价值计量且其变动计入当期损益\S*资产\S*': '以公允价值计量且其变动计入当期损益的资产',
       '(\S*|\S*\s*)衍生金融资产\S*': '衍生金融资产', '(\S*|\S*\s*)应收票据\S*': '应收票据', '(\S*|\S*\s*)应收账款\S*': '应收账款', '(\S*|\S*\s*)应收账款计提\S*': '应收账款计提',
       '(\S*|\S*\s*)预付款项\S*': '预付款项', '(\S*|\S*\s*)应收利息\S*': '应收利息', '(\S*|\S*\s*)应收股利\S*': '应收股利', '(\S*|\S*\s*)其他应收款\S*': '其他应收款',
       '(\S*|\S*\s*)其他应收款计提\S*': '其他应收款计提', '(\S*|\S*\s*)存货$': '存货', '(\S*|\S*\s*)划分为持有待售的资产\S*': '划分为持有待售的资产', '(\S*|\S*\s*)年内到期的非流动资产\S*': '一年内到期的非流动资产',
       '(\S*|\S*\s*)其他流动资产\S*': '其他流动资产', '(\S*|\S*\s*)可供出售金融资产\S*': '可供出售金融资产', '(\S*|\S*\s*)持有至到期投资\S*': '持有至到期投资',
       '(\S*|\S*\s*)长期应收款\S*': '长期应收款', '(\S*|\S*\s*)长期股权投资\S*': '长期股权投资', '(\S*|\S*\s*)投资性房地产\S*': '投资性房地产',
       '(\S*|\S*\s*)固定资产\S*': '固定资产', '(\S*|\S*\s*)在建工程\S*': '在建工程', '(\S*|\S*\s*)工程物资\S*': '工程物资', '(\S*|\S*\s*)固定资产清理\S*': '固定资产清理',
       '(\S*|\S*\s*)生物\S*资产\S*': '生产性生物资产', '(\S*|\S*\s*)油气资产\S*': '油气资产', '(\S*|\S*\s*)无形资产\S*': '无形资产', '(\S*|\S*\s*)开发支出\S*': '开发支出',
       '(\S*|\S*\s*)商誉\S*': '商誉', '(\S*|\S*\s*)长期待摊费用\S*': '长期待摊费用', '(\S*|\S*\s*)递延所得税资产产(负债)\S*': '递延所得税资产(负债)',
       '(\S*|\S*\s*)其他非流动资产\S*': '其他非流动资产', '(\S*|\S*\s*)短期借款\S*': '短期借款', '(\S*|\S*\s*)以公允价值计量且其变动计入当期损益\S*负债\S*': '以公允价值计量且其变动计入当期损益的负债',
       '(\S*|\S*\s*)衍生金融负债\S*': '衍生金融负债', '(\S*|\S*\s*)应付票据\S*': '应付票据', '(\S*|\S*\s*)应付账款\S*': '应付账款', '(\S*|\S*\s*)预收款项\S*': '预收款项',
       '(\S*|\S*\s*)应付职工薪酬\S*': '应付职工薪酬', '(\S*|\S*\s*)应交税费\S*': '应交税费', '(\S*|\S*\s*)应付利息\S*': '应付利息', '(\S*|\S*\s*)应付股利\S*': '应付股利',
       '(\S*|\S*\s*)其他应付款\S*': '其他应付款', '(\S*|\S*\s*)划分为持有待售的负债\S*': '划分为持有待售的负债', '(\S*|\S*\s*)年内到期的非流动负债\S*': '一年内到期的非流动负债',
       '(\S*|\S*\s*)其他流动负债\S*': '其他流动负债', '(\S*|\S*\s*)长期借款\S*': '长期借款', '(\S*|\S*\s*)应付债券\S*': '应付债券', '(\S*|\S*\s*)长期应付款\S*': '长期应付款',
       '(\S*|\S*\s*)长期应付职工薪酬\S*': '长期应付职工薪酬', '(\S*|\S*\s*)专项应付款\S*': '专项应付款', '(\S*|\S*\s*)预计负债\S*': '预计负债',
       '(\S*|\S*\s*)递延收益\S*': '递延收益', '(\S*|\S*\s*)其他非流动负债\S*': '其他非流动负债', '(\S*|\S*\s*)股本\S*': '股本', '(\S*|\S*\s*)其他权益工具\S*': '其他权益工具',
       '(\S*|\S*\s*)资本公积\S*': '资本公积', '(\S*|\S*\s*)库存股\S*': '库存股', '(\S*|\S*\s*)其他综合收益\S*': '其他综合收益', '(\S*|\S*\s*)专项储备\S*': '专项储备', '(\S*|\S*\s*)盈余公积\S*': '盈余公积',
       '(\S*|\S*\s*)未分配利润\S*': '未分配利润', '(\S*|\S*\s*)营业收入\S*营业成本\S*': '营业收入和营业成本', '(\S*|\S*\s*)营业税金\S*': '营业税金', '(\S*|\S*\s*)销售费用\S*': '销售费用',
       '(\S*|\S*\s*)管理费用\S*': '管理费用', '(\S*|\S*\s*)财务费用\S*': '财务费用', '(\S*|\S*\s*)资产减值损失\S*': '资产减值损失', '(\S*|\S*\s*)公允价值\S*收益\S*': '公允价值变动投资收益',
       '(\S*|\S*\s*)投资收益\S*': '投资收益', '(\S*|\S*\s*)营业外收入\S*': '营业外收入', '(\S*|\S*\s*)营业外支出\S*': '营业外支出', '(\S*|\S*\s*)所得税费用\S*': '所得税费用',
       '(\S*|\S*\s*)现金\S*等价物\S*': '现金和现金等价物的构成', '(\S*|\S*\s*)经营\S*租赁\S*': '经营租赁', '(\S*|\S*\s*)融资\S*租入\S*': '融资租入的固定资产', '(\S*|\S*\s*)坏账损失\S*': '资产减值损失'}

for root, Dir, files in os.walk(path):
    for filename in files:
        lap0 = datetime.datetime.now()    # lap 0
        wb = ox.load_workbook(os.path.join(root, filename), data_only=True)
        ws = wb.get_sheet_by_name("Table 1")

        if filename[0] == "0" or filename[0] == "2" or filename[0] == "3":
            Ticker = filename[0:6] + ".SZ"
        else:
            Ticker = filename[0:6] + ".SH"
        lap1 = datetime.datetime.now()    # lap 1
        print(Ticker + " R: ", lap1 - lap0)

        row_end = ws.max_row     # max_row
        col_end = ws.max_column   # max_column

        index_col = [ws.cell(row=i, column=col_end).value for i in range(1, row_end)]   # 标记需要提取的行数组成列表
        while None in index_col:
            index_col.remove(None)
        index_col_items = index_col[::3]
        index_col_frame = index_col[1::3] + index_col[2::3]
        index_col_frame.sort()
        items = [ws.cell(row=i, column=1).value for i in index_col_items]   # 按标记行提取标题 items

        for i in range(len(items)):
            for j in dic.keys():     # 对不规则的标题按dic key 和 value 规则化
                if trim_items_a(j, items[i]):
                    items[i] = dic[j]

        lap2 = datetime.datetime.now()   # lap 2
        print(Ticker, " Extr: ", lap2 - lap1)

        row = [[index_col_frame[x], index_col_frame[x + 1]] for x in range(len(index_col_frame) - 1)]
        row = row[0::2] # 删除重复项, [2,3]为重复项. [[1,2],[2,3],[3,4]...]
        [row[i].append(items[i]) for i in range(len(row))] # 将标题附加到row中
        # row_tmp = row.copy()     # 用于调试, 可删
        table = []
        # tmp0 = []    # 用于调试, 可删
        tmp1 = []
        tmp2 = []
        tmp3 = []
        tmp4 = []
        tmp5 = []
        lap3 = datetime.datetime.now()    # lap 3
        for i, j in enumerate(row):
            if isinstance(row[i][0], int) and isinstance(row[i][1], int):
                for m in range(row[i][0], row[i][1]):           # row[(0,1,2)]
                    l1 = [ws.cell(row=m, column=n).value for n in range(1, col_end - 1)]
                    l = clean_list(l1)
                    table.append(l)

                t_set = set()
                for p in range(len(table)):     # 辨认表格空白项目
                    for q in range(len(table[0])):
                        if table[p][q] is not '':
                            t_set.add(q)       # 删除表格空白项目
                            t_list = list(t_set)
                            t_list.sort()
                for x in range(len(table)):
                    var = [row[i][2]] + [Ticker] + [table[x][y] for y in t_list]    # 这样的形式: ['营业外收入', '002575.SZ', '非流动资产处置利得合计', 5248.08, '', 5248.08]
                    # tmp0.append(var)   # 用于调试, 可删

                    ws1 = wb1.get_sheet_by_name(var[0])
                    count = ws1.max_row + 1

                    if var[0] != '固定资产' and var[0] != '无形资产' and var[0] != '生产性生物资产' and var[0] != '油气资产' and var[
                        0] != '投资性房地产':     # 除LTA资产外, 开始按条件写入
                        ws1.append(var[1:])
                        # for n in range(1, len(var)):
                        #     ws1.cell(row=count, column=n, value=var[n])
                        count += 1
                    elif var[0] == '固定资产':  # 将LTA储存在tmp中
                        tmp1.append(var)
                    elif var[0] == '无形资产':
                        tmp2.append(var)
                    elif var[0] == '生产性生物资产':
                        tmp3.append(var)
                    elif var[0] == '油气资产':
                        tmp4.append(var)
                    elif var[0] == '投资性房地产':
                        tmp4.append(var)
                    else:
                        tmp5.append(var)

                table.clear()
        L=[tmp1,tmp2,tmp3,tmp4,tmp5]    # 从tmp中转换成dataframe, 记录在不同excel中
        for tmp in L:
            if len(tmp) != 0:          # 将没有数字的格子填充为np.nan
                for i in range(len(tmp)):
                    for j in range(len(tmp[i])):
                        if tmp[i][j] == '':
                            tmp[i][j] = np.nan
                if L.index(tmp)==0:         # 决定tmp是哪种资产
                    name='固定资产'
                elif L.index(tmp)==1:
                    name='无形资产'
                elif L.index(tmp)==2:
                    name='生产性生物资产'
                elif L.index(tmp)==3:
                    name='油气资产'
                elif L.index(tmp)==4:
                    name='投资性房地产'

                tmp = pd.DataFrame(tmp, columns=tmp[0]).drop(0).dropna(axis=1, how='all')   # 将第一行做column name,drop多出来的第一行, dropna
                tmp = tmp.set_index([tmp.columns[0], tmp.columns[1], tmp.columns[2]]).swaplevel(0,1)     # 将ticker, 项目, 具体科目 设置成 index, 两index互换
                tmp = tmp.stack().to_frame()                                                     # unpivot
                tmp = tmp.reset_index().values.tolist()                                          # reset index, 为了写入excel更方便以后pandas提取
                for row in tmp:
                    ws_LTA.append(row)
                wb_LTA.save('LTA.xlsx')
                # tmp.to_excel(name + ' ' + Ticker + '.xlsx')

        if len(tmp5)!=0:
            tmp5=pd.DataFrame(tmp5)
            tmp5.to_excel('tmp5.xlsx')
        row.clear()
        lap4 = datetime.datetime.now()   # lap 4
        print(Ticker, " W: ", lap4 - lap3)
        # LTA=pd.DataFrame(np.array(LTA))
        # LTA.to_excel('LTA.xlsx')
        # ____________________________________________________
        print(Ticker, " -> Done \n")
    wb1.save(filename="FS.xlsx")
print("*** Done Extracting ***")
