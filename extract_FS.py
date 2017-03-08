# -*- coding: utf-8 -*-
# by Shawn Gu

import openpyxl as ox
import numpy as np
import datetime, os, re
from openpyxl.styles import Font,colors

# _________________________ create a workbook for data storage_______________________

wb1 = ox.load_workbook('D:/pyscripts/FS.xlsx',data_only=True)  # 打开要录入的目的文件
ws1 = wb1.get_sheet_by_name('FS')

# ___________________________________________________________________________________
def clean_list(List):
    List = [str(List[i]).replace(" ", "") for i in range(len(List))]        # list value del space
    List = [str(List[i]).replace("\n", "") for i in range(len(List))]       # list value del \n
    List = [str(List[i]).replace("None", "") for i in range(len(List))]      # list value del "None"
    tmp = []
    for i in range(len(List)):
        try:
            tmp.append(float(List[i]))
        except ValueError:
            tmp.append(List[i])
    List = tmp
    return List

def trim_items_a(old, in_which):
    return re.match(old.encode('utf-8'), in_which.encode('utf-8')) is not None

def search_symbol(symbol,r,c):
    '''r=row, c=column in excel'''
    if ws1.cell(row=r, column=c).value!=None:
        ws1.cell(row=r,column=c).value=ws1.cell(row=r,column=c).value.replace(' ','')
        if re.search(symbol,ws1.cell(row=r,column=c).value) != None:
            return re.search(symbol,ws1.cell(row=r,column=c).value).span()[0]
        else:
            return 0
    else:
        return 0

# ___________________________________________________________________________________
row = []
count_1 = ws1.max_row + 1
count = ws1.max_row + 1

path = "C:/Users/Shawn Gu/Desktop/CN_annual_FS"             # 被提取文件所在地

dic={'(\S*|\S*\s*)合并\S*资产负债表\S*':'合并资产负债表','(\S*|\S*\s*)母\S*资产负债表\S*': '母公司资产负债表',
     '(\S*|\S*\s*)合并\S*利润表\S*': '合并利润表', '(\S*|\S*\s*)母\S*利润表\S*': '母公司利润表',
     '(\S*|\S*\s*)合并\S*现金流\S*': '合并现金流量表', '(\S*|\S*\s*)母\S*现金流\S*': '母公司现金流量表'}


acc=['以后不能重分类进损益的其他综合收益', '基本每股收益', '以后将重分类进损益的其他综合收益', '稀释每股收益', '权益法下在被投资单位以后将重分类进损益的其他综合收益中享有的份额',
     '重新计量设定受益计划净负债或净资产的变动', '可供出售金融资产公允价值变动损益', '权益法下在被投资单位不能重分类进损益的其他综合收益中享有的份额',
 '持有至到期投资重分类为可供出售金融资产损益', '现金流量套期损益的有效部分', '外币财务报表折算差额', '其他', '营业总收入', '营业收入', '一年内到期的非流动负债',
 '一年内到期的非流动资产', '一般风险准备', '每股收益：', '综合收益总额', '利润总额（亏损总额以“－”号填列）', '营业利润（亏损以“－”号填列）', '专项储备', '专项应付款',
 '买入返售金融资产', '营业利润（亏损以“－”号填列）', '营业总成本', '短期递延收益','其他综合收益的税后净额', '净利润（净亏损以“－”号填列）', '现金及现金等价物净增加额', '代理买卖证券款',
 '代理承销证券款', '以公允价值计量且其变动计入当期损益的金融负债', '以公允价值计量且其变动计入当期损益的金融资产', '保单红利支出', '保户储金及投资款净增加额', '保险合同准备金',
 '偿还债务支付的现金', '每股收益：', '其他综合收益的税后净额', '期末现金及现金等价物余额', '综合收益总额', '其中：优先股', '其中：子公司吸收少数股东投资收到的现金',
 '其中：子公司支付给少数股东的股利、利润', '其中：对联营企业和合营企业的投资收益', '其中：营业成本', '其中：营业收入', '其中：非流动资产处置利得', '其中：非流动资产处置损失',
 '其他应付款', '其他应收款', '其他权益工具', '其他流动负债', '其他流动资产', '其他综合收益', '其他非流动负债', '其他非流动资产', '减：库存股', '减：所得税费用',
 '减：营业外支出', '减：营业成本', '分保费用', '分配股利、利润或偿付利息支付的现金', '划分为持有待售的负债', '划分为持有待售的资产', '利息支出', '利息收入', '加：公允价值变动收益（损失以“－”号填列）',
 '加：期初现金及现金等价物余额', '加：营业外收入', '卖出回购金融资产款', '发放贷款及垫款', '发行债券收到的现金', '取得借款收到的现金', '取得子公司及其他营业单位支付的现金净额',
 '取得投资收益收到的现金', '可供出售金融资产', '向中央银行借款', '向中央银行借款净增加额', '向其他金融机构拆入资金净增加额', '吸收存款及同业存放', '吸收投资收到的现金',
 '商誉', '净利润（净亏损以“－”号填列）', '利润总额（亏损总额以“－”号填列）', '汇率变动对现金及现金等价物的影响', '回购业务资金净增加额', '固定资产', '固定资产清理',
 '在建工程', '处置以公允价值计量且其变动计入当期损益的金融资产净增加额', '处置固定资产、无形资产和其他长期资产收回的现金净额', '处置子公司及其他营业单位收到的现金净额',
 '存放中央银行和同业款项净增加额', '存货', '客户存款和同业存放款项净增加额', '客户贷款及垫款净增加额', '少数股东损益', '少数股东权益', '工程物资', '已赚保费', '应交税费',
 '应付债券', '应付分保账款', '应付利息', '应付手续费及佣金', '应付票据', '应付职工薪酬', '应付股利', '应付账款', '应收保费', '应收分保合同准备金', '应收分保账款', '应收利息',
 '应收票据', '应收股利', '应收账款', '开发支出', '归属于少数股东的其他综合收益的税后净额', '归属于少数股东的综合收益总额', '归属于母公司所有者权益合计', '归属于母公司所有者的净利润',
 '归属于母公司所有者的综合收益总额', '归属母公司所有者的其他综合收益的税后净额', '所有者权益合计', '手续费及佣金支出', '手续费及佣金收入', '投资性房地产', '投资支付的现金',
 '投资收益（损失以“－”号填列）', '投资活动产生的现金流量净额', '投资活动现金流入小计', '投资活动现金流出小计', '拆入资金', '拆入资金净增加额', '拆出资金', '持有至到期投资',
 '提取保险合同准备金净额', '支付保单红利的现金', '支付其他与投资活动有关的现金', '支付其他与筹资活动有关的现金', '支付其他与经营活动有关的现金', '支付利息、手续费及佣金的现金',
 '支付原保险合同赔付款项的现金', '支付的各项税费', '支付给职工以及为职工支付的现金', '收到其他与投资活动有关的现金', '收到其他与筹资活动有关的现金', '收到其他与经营活动有关的现金',
 '收到再保险业务现金净额', '收到原保险合同保费取得的现金', '收到的税费返还', '收取利息、手续费及佣金的现金', '收回投资收到的现金', '无形资产', '未分配利润', '永续债', '汇兑收益（损失以“－”号填列）',
 '油气资产', '流动负债合计', '流动资产合计', '生产性生物资产', '盈余公积', '短期借款', '筹资活动产生的现金流量净额', '筹资活动现金流入小计', '筹资活动现金流出小计', '管理费用',
 '经营活动产生的现金流量净额', '经营活动现金流入小计', '经营活动现金流出小计', '结算备付金', '股本', '能重分类进损益的其他综合收益中享有的份额', '营业税金及附加', '衍生金融负债',
 '衍生金融资产', '负债合计', '负债和所有者权益总计', '财务费用', '货币资金', '质押贷款净增加额', '购买商品、接受劳务支付的现金', '购建固定资产、无形资产和其他长期资产支付的现金',
 '资产减值损失', '资产总计', '资本公积', '赔付支出净额', '退保金', '递延所得税负债', '递延所得税资产', '递延收益', '销售商品、提供劳务收到的现金', '销售费用', '长期借款',
 '长期应付款', '长期应付职工薪酬', '长期应收款', '长期待摊费用', '长期股权投资', '非流动负债合计', '非流动资产合计', '预付款项', '预收款项', '预计负债']

ticker=[] # 在调整本期数字位置的时候用

for root, Dir, files in os.walk(path):
    for filename in files:
        lap0 = datetime.datetime.now()    # lap 0
        wb = ox.load_workbook(os.path.join(root, filename), data_only=True)
        ws = wb.get_sheet_by_name("Table 1")

        if filename[0] == "0" or filename[0] == "2" or filename[0] == "3":
            Ticker = filename[0:6] + ".SZ"
        else:
            Ticker = filename[0:6] + ".SH"
        lap1 = datetime.datetime.now()    # lap 1
        print(Ticker + " Read: ", lap1 - lap0)

        row_end = ws.max_row     # where None value row ends at
        col_end = ws.max_column

        index_col = [ws.cell(row=i, column=col_end).value for i in range(1, row_end-1)]
        while None in index_col:
            index_col.remove(None)
        index_col_items = index_col[::3]
        index_col_frame = index_col[1::3] + index_col[2::3]
        index_col_frame.sort()
        items = [ws.cell(row=i, column=1).value for i in index_col_items]
        for i in range(len(items)):
            for j in dic.keys():
                if trim_items_a(j, items[i]):
                    items[i] = dic[j]

        lap2 = datetime.datetime.now()   # lap 2
        print("Extraction: ", lap2 - lap1)

        row = [[index_col_frame[x], index_col_frame[x + 1]] for x in range(len(index_col_frame) - 1)]
        row = row[0::2]
        [row[i].append(items[i]) for i in range(len(row))]

        table = []
        type_ls=[]
#        tmp = [] # 调试以后删除
        for i, j in enumerate(row):
            if isinstance(row[i][0], int) and isinstance(row[i][1], int):
                for m in range(row[i][0], row[i][1]):           # row[(0,1,2)]
                    l1 = [ws.cell(row=m, column=n).value for n in range(1, col_end)]
                    l = clean_list(l1)
                    table.append(l)

                t_set = set()
                for p in range(len(table)):     # clean extracted table
                    for q in range(len(table[0])):
                        if table[p][q] is not '':
                            t_set.add(q)
                            t_list = list(t_set)
                            t_list.sort()
                lap3=datetime.datetime.now()

                for x in range(len(table)):
                    var = [Ticker] + [row[i][2]] + [table[x][y] for y in t_list]
                    ticker.append(var[0])
                    # tmp.append(var)    # 调试用, 可删
                    if re.match("流动资产：",var[2]) or re.match("非流动资产：",var[2]) or re.match("流动负债：",var[2]) or re.match("非流动负债：",var[2]) or re.match("所有者权益：", var[2]):
                        ws1.cell(row=count,column=4,value=var[2][:len(var[2])-1])
                    elif re.match("\S*经营活动\S*：", var[2]) or re.match('(\S*|\S*\s*)投资活动\S*：',var[2]) or re.match('(\S*|\S*\s*)筹资活动\S*：',var[2]):
                        ws1.cell(row=count,column=4,value=var[2][2:len(var[2])-1])
                    elif re.match('(\S*|\S*\s*)利润表',var[1]):
                        ws1.cell(row=count,column=4,value=var[1])
                        for col in range(0, 3):
                            ws1.cell(row=count,column=col+1, value=var[col])
                        for col in range(3, len(var)):
                            ws1.cell(row=count,column=col+2, value=var[col])
                        count += 1
                    else:
                        for col in range(0, 3):
                            ws1.cell(row=count, column=col+1, value=var[col])
                        for col in range(3, len(var)):
                            ws1.cell(row=count, column=col+2, value=var[col])
                        count += 1
                table.clear()
        row.clear()
        print(Ticker, "Done _______ \n")

for i in range(count_1,count):
    if type(ws1.cell(row=i,column=5).value)!=float and type(ws1.cell(row=i,column=5).value)!= int and ws1.cell(row=i, column=5).value!='':
        if re.search('）',ws1.cell(row=i, column=5).value)!=None and re.search('（',ws1.cell(row=i, column=5).value)==None: # 在2014年会存在 '）' (包括附注标识) 抛出表格, 将其替换成 np.NaN
            ws1.cell(row=i,column=5, value=ws1.cell(row=i,column=6).value)
            ws1.cell(row=i,column=6, value=0)
        elif ws1.cell(row=i,column=5).value!='）':
            ws1.cell(row=i,column=5).value=''
    # elif type(ws1.cell(row=i,column=5).value)!=float and type(ws1.cell(row=i,column=5).value)!= int and ws1.cell(row=i, column=5).value!='' and ws1.cell(row=i,column=5).value!='）':
    #     ws1.cell(row=i,column=5).value=''
ticker=set(ticker)

for t in ticker:
    ls_empty = []
    for i in range(1,ws1.max_row+1):
        if ws1.cell(row=i,column=1).value==t:
            ls_empty.append(ws1.cell(row=i, column=5).value)
    if len(set(ls_empty))==1:
        for i in range(1,ws1.max_row+1):
            if ws1.cell(row=i,column=1).value==t:
                ws1.cell(row=i,column=5).value=ws1.cell(row=i,column=6).value
                ws1.cell(row=i,column=6).value=''

# __________________________________________________________________________________________________
count_2 = ws1.max_row

for i in range(2, count_2+1):
    A=search_symbol('）', i, 3)
    B=search_symbol('、', i, 3)
    C=search_symbol('\.', i, 3)
    x=[A,B,C]
    while 0 in x:
        x.remove(0)         # 将0排除在外, 因为0代表无特殊符号
    if x!=[]:
        if min(x)<4:        # 这些特殊符号应该不超过 4 这个位置
            ws1.cell(row=i, column=8, value=ws1.cell(row=i,column=3).value)               # 将原述科目一到H列, 可drop
            ws1.cell(row=i, column=3,value=ws1.cell(row=i, column=3).value[min(x)+1:])    # 将有特殊符号的直接替换成标准格式

    if ws1.cell(row=i,column=3).value not in acc:                   # 查看这一格是否是在acc中的标准格式
        ws1.cell(row=i,column=3).font=Font(color=colors.RED)        # 不是标准格式的 标红

wb1.save(filename="D:/pyscripts/FS.xlsx")
print(" *** Done Extracting *** ")